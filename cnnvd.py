#coding:utf-8
#Author: fa1lr4in
from openpyxl import Workbook
import requests
from bs4 import BeautifulSoup as BS
import sys,re,time
import calendar
import re,os

def mkdir():
	cwd = os.getcwd()
	path = cwd + "\\cnnvd"
	path=path.strip()
	path=path.rstrip("\\")
	isExists=os.path.exists(path)

	# 判断结果
	if not isExists:
		os.makedirs(path) 
		print ('create dir success')
		return True
	else:
		print ('dir is exist')
		return False

def cnnvd():
	mkdir()
	wb = Workbook()
	vuln_name = input("请输入要搜集的组件名称：")
	print('你要搜集的组件名称为' + vuln_name)
	url = 'http://www.cnnvd.org.cn/web/vulnerability/queryLds.tag'
	data = 'CSRFToken=&cvHazardRating=&cvVultype=&qstartdateXq=&cvUsedStyle=&cvCnnvdUpdatedateXq=&cpvendor=&relLdKey=&hotLd=&isArea=&qcvCname=' + vuln_name + '&qcvCnnvdid=CNNVD%E6%88%96CVE%E7%BC%96%E5%8F%B7&qstartdate=&qenddate='
	headers = {
			'Host': 'www.cnnvd.org.cn',
			'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64; rv:79.0) Gecko/20100101 Firefox/79.0',
			'Accept': 'text/html,application/xhtml+xml,application/xml;q=0.9,image/webp,*/*;q=0.8',
			'Accept-Language': 'zh-CN,zh;q=0.8,zh-TW;q=0.7,zh-HK;q=0.5,en-US;q=0.3,en;q=0.2',
			'Accept-Encoding': 'gzip, deflate',
			'Content-Type': 'application/x-www-form-urlencoded',
			'Content-Length': '205',
			'Origin': 'http://www.cnnvd.org.cn',
			'Connection': 'close',
			'Referer': 'http://www.cnnvd.org.cn/web/vulnerability/queryLds.tag',
			'Cookie': 'SESSION=4a66e9df-3567-4a06-8239-6489f9cc5770',
			'Upgrade-Insecure-Requests': '1'
		  }
	resp = requests.post(url,data=data,headers=headers)
	#print(resp.text)
	soup = BS(resp.text,'lxml')
	sheet_nvd = wb.create_sheet(index=0, title="cnvd") 
	sheets = wb.sheetnames
	ws = wb[sheets[0]]
	rows = ws.rows
	columns = ws.columns
	ws.cell(row=1, column=1).value = "CVE"
	ws.cell(row=1, column=2).value = "CNNVD"
	ws.cell(row=1, column=3).value = "漏洞描述"
	ws.cell(row=1, column=4).value = "漏洞公开日期"
	#get vulns_num 
	tmp_html = soup.find_all(name='div',attrs={'class':'page'})
	tmp_soup = BS(str(tmp_html),'lxml')
	tmp = tmp_soup.find(name='a',attrs={'onmouse':''})
	tmp_pattern = re.compile(r'\d{1,10}')
	vulns_num = int(tmp_pattern.search(tmp.string).group())
	print(vulns_num)
	#get vulns
	loop_time = 0
	for arg in range(0,vulns_num,10):
		loop_time = loop_time + 1
		url = 'http://www.cnnvd.org.cn/web/vulnerability/queryLds.tag?pageno=' + str(loop_time) + '&repairLd='
		data = 'CSRFToken=&cvHazardRating=&cvVultype=&qstartdateXq=&cvUsedStyle=&cvCnnvdUpdatedateXq=&cpvendor=&relLdKey=&hotLd=&isArea=&qcvCname=' + vuln_name + '&qcvCnnvdid=CNNVD%E6%88%96CVE%E7%BC%96%E5%8F%B7&qstartdate=&qenddate='
		headers = {
				'Host': 'www.cnnvd.org.cn',
				'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64; rv:79.0) Gecko/20100101 Firefox/79.0',
				'Accept': 'text/html,application/xhtml+xml,application/xml;q=0.9,image/webp,*/*;q=0.8',
				'Accept-Language': 'zh-CN,zh;q=0.8,zh-TW;q=0.7,zh-HK;q=0.5,en-US;q=0.3,en;q=0.2',
				'Accept-Encoding': 'gzip, deflate',
				'Content-Type': 'application/x-www-form-urlencoded',
				'Content-Length': '205',
				'Origin': 'http://www.cnnvd.org.cn',
				'Connection': 'close',
				'Referer': 'http://www.cnnvd.org.cn/web/vulnerability/queryLds.tag',
				'Cookie': 'SESSION=4a66e9df-3567-4a06-8239-6489f9cc5770',
				'Upgrade-Insecure-Requests': '1'
				}
		resp = requests.post(url,data=data,headers=headers)
		resp_text = resp.text
		#战术sleep
		time.sleep(3)
		soup = BS(resp_text,'lxml')
		tmp_pattern = re.compile('CNNVD-.*?target="_blank">')
		vulns_cnnvd = tmp_pattern.findall(resp_text)
		for vuln in vulns_cnnvd:
			tmp_pattern1 = re.compile('CNNVD-\d{1,9}-\d{1,9}')
			vuln = tmp_pattern1.findall(vuln)
			cnnvd_number = vuln[0]
			#print(cnnvd_number)
		#######################		get date 		#############################	
		current_row = arg + 2
		vulns_date = soup.find_all(name='div',attrs={'class':'fr','style':'margin-top:8px;'})
		for vuln in vulns_date:
			vuln = str(vuln)
			tmp_pattern2 = re.compile('\d{3,5}-\d{1,3}-\d{1,3}')
			vuln = tmp_pattern2.findall(vuln)
			cnnvd_date = vuln[0]
			ws.cell(row=current_row, column=4).value = cnnvd_date
			current_row = current_row + 1


		current_row = arg + 2
		for vuln in vulns_cnnvd:
			tmp_pattern1 = re.compile('CNNVD-\d{1,9}-\d{1,9}')
			vuln = tmp_pattern1.findall(vuln)
			vuln = vuln[0]
			#add cve number,bid,title
			url = 'http://www.cnnvd.org.cn/web/xxk/ldxqById.tag?CNNVD=' + vuln
			resp = requests.get(url)
			#print(resp)
			soup = BS(resp.text,'lxml')
			vulns_cve = soup.find_all(name='a',attrs={'href':re.compile('cve.mitre.org')})
			for cve in vulns_cve:
				print(cve.string)
				ws.cell(row=current_row, column=1).value = cve.string
			#print(soup.title.string)

			tmp_html = soup.find_all(name='div',attrs={'class':'detail_xq w770'})
			tmp_soup = BS(str(tmp_html),'lxml')
			tmp = tmp_soup.find(name='h2')
			print(tmp.string)
			ws.cell(row=current_row, column=3).value = tmp.string
			#end add cve number,bid,title
			#add cnvd
			ws.cell(row=current_row, column=2).value = vuln
			current_row = current_row + 1


	wb.save('cnnvd\\' + vuln_name + '.xlsx')


if __name__ == "__main__":
	cnnvd()
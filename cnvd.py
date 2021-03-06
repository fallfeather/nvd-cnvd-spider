#coding:utf-8
#Author: fa1lr4in
from openpyxl import Workbook
import requests
from bs4 import BeautifulSoup as BS
import sys,re,time
import calendar
import re,os

def mkdir():
	cwd = os.getcwd()
	path = cwd + "\\cnvd"
	path=path.strip()
	path=path.rstrip("\\")
	isExists=os.path.exists(path)

	# 判断结果
	if not isExists:
		os.makedirs(path) 
		print ('create dir success')
		return True
	else:
		print ('dir is exist')
		return False

def change_date(date):
	year_tmp = str(re.findall('\d{4}',date))[2:-2]
	mouth_tmp = str(list(calendar.month_name).index(str(re.findall('[A-Za-z]{3,15}',date))[2:-2]))
	day_tmp = str(re.findall('\d{2}?',date))[2:4]
	return year_tmp + '/' + mouth_tmp + '/' + day_tmp

def cnvd():
	mkdir()
	wb = Workbook()
	vuln_name = input("请输入要搜集的组件名称：")
	print('你要搜集的组件名称为' + vuln_name)
	url = 'https://www.cnvd.org.cn/flaw/list.htm?flag=true'
	data = 'number=%E8%AF%B7%E8%BE%93%E5%85%A5%E7%B2%BE%E7%A1%AE%E7%BC%96%E5%8F%B7&startDate=&endDate=&field=&order=&flag=true&keyword=' + vuln_name + '&condition=1&keywordFlag=0&cnvdId=&cnvdIdFlag=0&baseinfoBeanbeginTime=&baseinfoBeanendTime=&baseinfoBeanFlag=0&refenceInfo=&referenceScope=-1&manufacturerId=-1&categoryId=-1&editionId=-1&causeIdStr=&threadIdStr=&serverityIdStr=&positionIdStr=&max=20&offset=0'
	headers = {
				'Host': 'www.cnvd.org.cn',
				'User-Agent': 'Mozilla/5.0 (X11; Linux x86_64; rv:68.0) Gecko/20100101 Firefox/68.0',
				'Accept': 'text/html,application/xhtml+xml,application/xml;q=0.9,*/*;q=0.8',
				'Accept-Language': 'zh-CN,zh;q=0.8,zh-TW;q=0.7,zh-HK;q=0.5,en-US;q=0.3,en;q=0.2',
				'Accept-Encoding': 'gzip, deflate',
				'Referer': 'https://www.cnvd.org.cn/flaw/list.htm?flag=true',
				'Content-Type': 'application/x-www-form-urlencoded',
				'Content-Length': '256',
				'Connection': 'close',
				'Cookie': '__jsluid_s=dc8ce6ec4bdcf4b0ea067886e4e6359e; JSESSIONID=48DC8FDD630411A285DEA4A33DDE54B9',
				'Upgrade-Insecure-Requests': '1'
			  }
	resp = requests.post(url,data=data,headers=headers)
	soup = BS(resp.text,'lxml')
	sheet_nvd = wb.create_sheet(index=0, title="cnvd") 
	sheets = wb.sheetnames
	ws = wb[sheets[0]]
	rows = ws.rows
	columns = ws.columns
	ws.cell(row=1, column=1).value = "CVE"
	ws.cell(row=1, column=2).value = "bugtraq-id"
	ws.cell(row=1, column=3).value = "CNVD"
	ws.cell(row=1, column=4).value = "漏洞描述"
	#ws.cell(row=1, column=3).value = "危险等级"
	#ws.cell(row=1, column=4).value = "CVSS2"
	ws.cell(row=1, column=5).value = "漏洞公开日期"
	#get vulns_num 
	tmp_html = soup.find_all(name='div',attrs={'class':'pages clearfix'})
	tmp_soup = BS(str(tmp_html),'lxml')
	tmp = tmp_soup.find_all(name='span')
	try:
		tmp1 = tmp[2]
	except:
		tmp1 = tmp[1]
	tmp_pattern = re.compile(r'\d{1,10}')
	try:
		vulns_num = int(tmp_pattern.search(tmp1.string).group())
	except:
		tmp1 = tmp[0]
		vulns_num = int(tmp_pattern.search(tmp1.string).group())
	#get vulns
	for arg in range(0,vulns_num,20):
		url = 'https://www.cnvd.org.cn/flaw/list.htm?flag=true'
		data = 'number=%E8%AF%B7%E8%BE%93%E5%85%A5%E7%B2%BE%E7%A1%AE%E7%BC%96%E5%8F%B7&startDate=&endDate=&field=&order=&flag=true&keyword=' + vuln_name + '&condition=1&keywordFlag=0&cnvdId=&cnvdIdFlag=0&baseinfoBeanbeginTime=&baseinfoBeanendTime=&baseinfoBeanFlag=0&refenceInfo=&referenceScope=-1&manufacturerId=-1&categoryId=-1&editionId=-1&causeIdStr=&threadIdStr=&serverityIdStr=&positionIdStr=&max=20&offset=' + str(arg)
		headers = {
				'Host': 'www.cnvd.org.cn',
				'User-Agent': 'Mozilla/5.0 (X11; Linux x86_64; rv:68.0) Gecko/20100101 Firefox/68.0',
				'Accept': 'text/html,application/xhtml+xml,application/xml;q=0.9,*/*;q=0.8',
				'Accept-Language': 'zh-CN,zh;q=0.8,zh-TW;q=0.7,zh-HK;q=0.5,en-US;q=0.3,en;q=0.2',
				'Accept-Encoding': 'gzip, deflate',
				'Referer': 'https://www.cnvd.org.cn/flaw/list.htm?flag=true',
				'Content-Type': 'application/x-www-form-urlencoded',
				'Content-Length': '256',
				'Connection': 'close',
				'Cookie': '__jsluid_s=dc8ce6ec4bdcf4b0ea067886e4e6359e; JSESSIONID=48DC8FDD630411A285DEA4A33DDE54B9',
				'Upgrade-Insecure-Requests': '1'
			  		}
		resp = requests.post(url,data=data,headers=headers)
		resp_text = resp.text
		#战术sleep
		time.sleep(3)
		soup = BS(resp.text,'lxml')
		vulns_tbody = soup.tbody
		soup_vulns = BS(str(vulns_tbody),'lxml')
		tmp_pattern = re.compile('/flaw/show/CNVD-.*?"\n')
		vulns_cnvd = tmp_pattern.findall(resp_text)
		#print(type(vulns_cnvd))
		#vulns_cnvd = vulns_cnvd[11:]
		#vulns_cnvd = soup_vulns.find_all(name='a',attrs={'href':re.compile('/flaw/show/CNVD-.*?')})
		current_row = arg + 2
		vulns_cvss2 = soup_vulns.find_all(name='td',attrs={'width':re.compile('13.*?')})
		for vuln in vulns_cvss2:
			if(vuln.string != None):
				print(vuln.string)
				#print(type(vuln.string))
				ws.cell(row=current_row, column=5).value = vuln.string
				current_row = current_row + 1


		current_row = arg + 2
		for vuln in vulns_cnvd:
			vuln = vuln[11:-2]
			print(vuln)
			#add cve number,bid,title
			url = 'https://www.cnvd.org.cn/flaw/show/' + vuln
			resp = requests.get(url)
			soup = BS(resp.text,'lxml')
			vulns_tbody = soup.tbody
			soup_vulns = BS(str(vulns_tbody),'lxml')
			vulns_cve = soup_vulns.find_all(name='a',attrs={'href':re.compile('cve.mitre.org')})
			for cve in vulns_cve:
				print(cve.string)
				ws.cell(row=current_row, column=1).value = cve.string
			vulns_bid = soup_vulns.find_all(name='a',attrs={'href':re.compile('www.securityfocus.com')})
			for bid in vulns_bid:
				print(bid.string)
				ws.cell(row=current_row, column=2).value = bid.string

			tmp_html = soup.find_all(name='div',attrs={'class':'blkContainerSblk'})
			tmp_soup = BS(str(tmp_html),'lxml')
			tmp = tmp_soup.find(name='h1')
			print(tmp.string)
			ws.cell(row=current_row, column=4).value = tmp.string
			#end add cve number,bid,title
			#add cnvd
			ws.cell(row=current_row, column=3).value = vuln
			current_row = current_row + 1
		# current_row = arg + 2
		# vulns_desc = soup_vulns.find_all(name='a',attrs={'title':re.compile('.*?')})
		# for vuln in vulns_desc:
		# 	print((vuln.string))
		# 	ws.cell(row=current_row, column=4).value = vuln.string
		# 	current_row = current_row + 1

	wb.save('cnvd\\' + vuln_name + '.xlsx')

#def seebug():
	

if __name__ == '__main__':
	cnvd()
#coding:utf-8
#Author: fa1lr4in
#核心思路： 发送一个请求，通过bs4处理resp数据，进行数据提取
from openpyxl import Workbook
import requests
from bs4 import BeautifulSoup as BS
import sys,re,time
import calendar
import execjs  
import urllib.request
import os

def mkdir():
	cwd = os.getcwd()
	path = cwd + "\\nvd"
	path=path.strip()
	path=path.rstrip("\\")
	isExists=os.path.exists(path)

	# 判断结果
	if not isExists:
		os.makedirs(path) 
		print ('create dir success')
		return True
	else:
		print ('dir is exist')
		return False

#google翻译接口：start
class Py4Js():
	def __init__(self):
		self.ctx = execjs.compile("""
		function TL(a) {
		var k = "";
		var b = 406644;
		var b1 = 3293161072;
		 
		var jd = "."; 
		var $b = "+-a^+6"; 
		var Zb = "+-3^+b+-f"; 

		for (var e = [], f = 0, g = 0; g < a.length; g++) { 
			var m = a.charCodeAt(g); 
			128 > m ? e[f++] = m : (2048 > m ? e[f++] = m >> 6 | 192 : (55296 == (m & 64512) && g + 1 < a.length && 56320 == (a.charCodeAt(g + 1) & 64512) ? (m = 65536 + ((m & 1023) << 10) + (a.charCodeAt(++g) & 1023), 
			e[f++] = m >> 18 | 240, 
			e[f++] = m >> 12 & 63 | 128) : e[f++] = m >> 12 | 224, 
			e[f++] = m >> 6 & 63 | 128), 
			e[f++] = m & 63 | 128) 
		} 
		a = b; 
		for (f = 0; f < e.length; f++) a += e[f], 
		a = RL(a, $b); 
		a = RL(a, Zb); 
		a ^= b1 || 0; 
		0 > a && (a = (a & 2147483647) + 2147483648); 
		a %= 1E6; 
		return a.toString() + jd + (a ^ b) 
	}; 
	 
	function RL(a, b) { 
		var t = "a"; 
		var Yb = "+"; 
		for (var c = 0; c < b.length - 2; c += 3) { 
			var d = b.charAt(c + 2), 
			d = d >= t ? d.charCodeAt(0) - 87 : Number(d), 
			d = b.charAt(c + 1) == Yb ? a >>> d: a << d; 
			a = b.charAt(c) == Yb ? a + d & 4294967295 : a ^ d 
		} 
		return a 
	} 
	""")

	def getTk(self,text):  
		return self.ctx.call("TL",text)
 
def open_url(url):
	headers = {'User-Agent': 'Mozilla/5.0 (Windows NT 6.1; WOW64; rv:23.0) Gecko/20100101 Firefox/23.0'}
	req = urllib.request.Request(url=url, headers=headers)
	response = urllib.request.urlopen(req)
	data = response.read().decode('utf-8')
	return data
 
def translate(content, tk):
	global rules_lines
	if len(content) > 4891:
		print("翻译的长度超过限制！！！")
		return

	content = urllib.parse.quote(content)

	url = "http://translate.google.cn/translate_a/single?client=t" + "&sl=en&tl=zh-CN&hl=zh-CN&dt=at&dt=bd&dt=ex&dt=ld&dt=md&dt=qca" + "&dt=rw&dt=rm&dt=ss&dt=t&ie=UTF-8&oe=UTF-8&clearbtn=1&otf=1&pc=1" + "&srcrom=0&ssel=0&tsel=0&kc=2&tk=%s&q=%s" % (
	tk, content)

	# 返回值是一个多层嵌套列表的字符串形式，解析起来还相当费劲，写了几个正则，发现也很不理想，
	# 后来感觉，使用正则简直就是把简单的事情复杂化，这里直接切片就Ok了
	result = open_url(url)

	end = result.find("\",")
	return result[4:end]
#google翻译接口：end

#利用正则匹配年月日，改变日期显示格式
def change_date(date):
	year_tmp = str(re.findall('\d{4}',date))[2:-2]
	mouth_tmp = str(list(calendar.month_name).index(str(re.findall('[A-Za-z]{3,15}',date))[2:-2]))
	day_tmp = str(re.findall('\d{2}?',date))[2:4]
	return year_tmp + '/' + mouth_tmp + '/' + day_tmp

#nvd主爬虫，将表格里面的东西除了漏洞中文描述全部填写上
def nvd():
	mkdir()
	wb = Workbook()
	vuln_name = input("请输入要搜集的组件名称：")
	print('你要搜集的组件名称为' + vuln_name)
	url = 'https://nvd.nist.gov/vuln/search/results?form_type=Advanced&results_type=overview&query=' + vuln_name + '&search_type=all&startIndex=0'
	resp = requests.get(url)
	soup = BS(resp.text,'lxml')

	#open excel
	sheet_nvd = wb.create_sheet(index=0, title="nvd") 
	sheets = wb.sheetnames
	ws = wb[sheets[0]]
	rows = ws.rows
	columns = ws.columns
	ws.cell(row=1, column=1).value = "CVE"
	ws.cell(row=1, column=2).value = "漏洞描述"
	ws.cell(row=1, column=3).value = "CVSS3"
	ws.cell(row=1, column=4).value = "CVSS2"
	ws.cell(row=1, column=5).value = "漏洞公开日期"
	ws.cell(row=1, column=6).value = "漏洞中文描述"
	#get vulns_num 
	tmp_html = soup.find_all(name='div',attrs={'class':'col-sm-12 col-lg-3'})
	tmp_soup = BS(str(tmp_html),'lxml')
	tmp = tmp_soup.find_all(name='strong',attrs={'data-testid':'vuln-matching-records-count'})
	for i in tmp:
		vulns_num = int(i.string)
	#get vulns
	for arg in range(0,vulns_num,20):
		desc = ""
		url = 'https://nvd.nist.gov/vuln/search/results?form_type=Advanced&results_type=overview&query=' + vuln_name + '&search_type=all&startIndex=' + str(arg)
		resp = requests.get(url)
		#战术sleep
		time.sleep(3)
		soup = BS(resp.text,'lxml')
		vulns_tbody = soup.tbody
		soup_vulns = BS(str(vulns_tbody),'lxml')
		vulns_cve = soup_vulns.find_all(name='a',attrs={'href':re.compile('/vuln/detail/CVE-.*?')})
		current_row = arg + 2
		for vuln in vulns_cve:
			print(vuln.string)
			ws.cell(row=current_row, column=1).value = vuln.string
			current_row = current_row + 1
		current_row = arg + 2
		vulns_desc = soup_vulns.find_all(name='p',attrs={'data-testid':re.compile('vuln-summary.*?')})
		for vuln in vulns_desc:
			print(vuln.string)
			desc = vuln.string
			ws.cell(row=current_row, column=2).value = vuln.string
			#中文描述start，翻译的太nc了，考虑要不要砍掉
			js = Py4Js()
			tk = js.getTk(desc)
			vuln_tmp = translate(desc, tk)
			print(vuln_tmp)
			ws.cell(row=current_row, column=6).value = vuln_tmp
			#中文描述功能end
			current_row = current_row + 1
		current_row = arg + 2
		vulns_cvss3 = soup_vulns.find_all(name='a',attrs={'data-testid':re.compile('vuln-cvss3-link.*?')})
		for vuln in vulns_cvss3:
			print(vuln.string)
			ws.cell(row=current_row, column=3).value = vuln.string
			current_row = current_row + 1
		current_row = arg + 2
		vulns_cvss2 = soup_vulns.find_all(name='a',attrs={'data-testid':re.compile('vuln-cvss2-link.*?')})
		for vuln in vulns_cvss2:
			print(vuln.string)
			ws.cell(row=current_row, column=4).value = vuln.string
			current_row = current_row + 1
		current_row = arg + 2
		vulns_date = soup_vulns.find_all(name='span',attrs={'data-testid':re.compile('vuln-published-on.*?')})
		for vuln in vulns_date:
			vuln_tmp = str(re.findall(".*;",vuln.string))
			vuln_tmp = change_date(str(vuln_tmp))
			print(vuln_tmp)
			ws.cell(row=current_row, column=5).value = vuln_tmp
			current_row = current_row + 1
	wb.save('nvd\\' + vuln_name + '.xlsx')
	


if __name__ == '__main__':
	nvd()